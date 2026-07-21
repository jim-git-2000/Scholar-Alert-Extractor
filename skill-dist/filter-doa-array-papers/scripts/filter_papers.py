#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import os
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font

VERSION = 2
SOURCE_SHEET = "Papers"
SOURCE_HEADERS = [
    "title",
    "authors",
    "year",
    "publication",
    "doi",
    "ieee_document_id",
    "paper_url",
    "pdf_url",
    "snippet",
    "sources",
    "alert_names",
    "first_seen_at",
    "last_seen_at",
    "seen_count",
    "dedup_key",
]
SEMANTIC_FIELDS = [
    "title",
    "authors",
    "year",
    "publication",
    "doi",
    "ieee_document_id",
    "paper_url",
    "pdf_url",
    "snippet",
    "sources",
    "alert_names",
]
OUTPUT_META_HEADERS = [
    "source_row",
    "relevance_scope",
    "relevance_reason",
    "matched_topics",
    "evidence_basis",
    "classified_at",
]
DECISIONS = {"relevant", "review", "excluded"}
SCOPES = {
    "core-doa",
    "array-signal-processing",
    "array-design-calibration",
    "joint-spatial-estimation",
    "borderline",
    "out-of-scope",
}
EVIDENCE_BASES = {
    "title",
    "title-and-snippet",
    "metadata-insufficient",
}


class FilterError(RuntimeError):
    pass


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def json_bytes(value: Any) -> bytes:
    return json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def fingerprint(values: dict[str, Any]) -> str:
    semantic = {field: json_value(values.get(field)) for field in SEMANTIC_FIELDS}
    return hashlib.sha256(json_bytes(semantic)).hexdigest()


def record_id(values: dict[str, Any], row_number: int) -> str:
    key = str(values.get("dedup_key") or "").strip()
    if key:
        return key
    fallback = {
        "row": row_number,
        "title": json_value(values.get("title")),
        "authors": json_value(values.get("authors")),
        "year": json_value(values.get("year")),
    }
    return f"row-fallback:{row_number}:{hashlib.sha256(json_bytes(fallback)).hexdigest()[:16]}"


def read_source(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        raise FilterError(f"源 Excel 不存在: {path}")
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise FilterError(f"无法读取源 Excel: {exc}") from exc
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise FilterError(f"源 Excel 缺少 {SOURCE_SHEET} 工作表")
        sheet = workbook[SOURCE_SHEET]
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, ())
        actual = list(header_row[: len(SOURCE_HEADERS)])
        if actual != SOURCE_HEADERS:
            raise FilterError("源 Excel 表头与 Scholar Alert Extractor 固定字段不一致")
        records: list[dict[str, Any]] = []
        seen_ids: set[str] = set()
        for row_number, row in enumerate(rows, start=2):
            values = dict(zip(SOURCE_HEADERS, row[: len(SOURCE_HEADERS)], strict=False))
            if not str(values.get("title") or "").strip():
                continue
            item_id = record_id(values, row_number)
            if item_id in seen_ids:
                raise FilterError(f"源 Excel 存在重复 dedup_key: {item_id}")
            seen_ids.add(item_id)
            records.append(
                {
                    "id": item_id,
                    "fingerprint": fingerprint(values),
                    "source_row": row_number,
                    "values": values,
                }
            )
        return records
    finally:
        workbook.close()


def load_json(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise FilterError(f"文件不存在: {path}") from exc
    except (OSError, json.JSONDecodeError) as exc:
        raise FilterError(f"无法读取 JSON {path}: {exc}") from exc
    if not isinstance(value, dict):
        raise FilterError(f"JSON 顶层必须是对象: {path}")
    return value


def load_state(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {
            "version": VERSION,
            "last_source_row": 1,
            "checkpoint_digest": prefix_digest([], 1),
            "records": {},
        }
    state = load_json(path)
    if (
        state.get("version") != VERSION
        or not isinstance(state.get("last_source_row"), int)
        or not isinstance(state.get("checkpoint_digest"), str)
        or not isinstance(state.get("records"), dict)
    ):
        raise FilterError(f"不支持或损坏的增量状态: {path}")
    return state


def atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            prefix=f".{path.name}-",
            suffix=".tmp",
            dir=path.parent,
            delete=False,
        ) as handle:
            temporary = Path(handle.name)
            json.dump(payload, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
        temporary = None
    finally:
        if temporary is not None:
            temporary.unlink(missing_ok=True)


def prefix_digest(records: list[dict[str, Any]], last_source_row: int) -> str:
    anchors = [
        {
            "source_row": record["source_row"],
            "title": json_value(record["values"].get("title")),
            "first_seen_at": json_value(record["values"].get("first_seen_at")),
        }
        for record in records
        if record["source_row"] <= last_source_row
    ]
    return hashlib.sha256(json_bytes(anchors)).hexdigest()


def validate_checkpoint(records: list[dict[str, Any]], state: dict[str, Any]) -> None:
    last_source_row = state["last_source_row"]
    if last_source_row > 1 and not any(
        record["source_row"] == last_source_row for record in records
    ):
        raise FilterError(
            "源 Excel 的历史行被删除或移动，无法安全继续增量筛选"
        )
    if prefix_digest(records, last_source_row) != state["checkpoint_digest"]:
        raise FilterError(
            "源 Excel 在 checkpoint 之前被重排或改名，无法安全继续增量筛选"
        )


def current_counts(records: list[dict[str, Any]], state: dict[str, Any]) -> dict[str, int]:
    counts = {"relevant": 0, "review": 0, "excluded": 0, "new": 0}
    for saved in state["records"].values():
        if not isinstance(saved, dict):
            continue
        decision = saved.get("decision")
        if decision in DECISIONS:
            counts[decision] += 1
    counts["new"] = sum(
        record["source_row"] > state["last_source_row"] for record in records
    )
    return counts


def pending_view(record: dict[str, Any]) -> dict[str, Any]:
    values = record["values"]
    return {
        "id": record["id"],
        "fingerprint": record["fingerprint"],
        "source_row": record["source_row"],
        **{field: json_value(values.get(field)) for field in SEMANTIC_FIELDS},
    }


def batch_token(items: list[dict[str, Any]], source_snapshot: str) -> str:
    material = {
        "source_snapshot": source_snapshot,
        "items": [{"id": item["id"], "fingerprint": item["fingerprint"]} for item in items],
    }
    return hashlib.sha256(json_bytes(material)).hexdigest()


def command_prepare(args: argparse.Namespace) -> None:
    source = args.source.resolve()
    state_path = args.state.resolve()
    output = args.output.resolve()
    pending_path = args.pending.resolve()
    decisions_path = args.decisions.resolve()
    records = read_source(source)
    state = load_state(state_path)
    validate_checkpoint(records, state)
    source_snapshot = sha256_file(source)
    all_new = [
        record for record in records if record["source_row"] > state["last_source_row"]
    ]
    selected = all_new[: args.limit]
    visible_items = [pending_view(record) for record in selected]
    token = batch_token(visible_items, source_snapshot)
    pending_payload = {
        "version": VERSION,
        "batch_token": token,
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source": str(source),
        "source_snapshot": source_snapshot,
        "checkpoint_before": state["last_source_row"],
        "checkpoint_digest_before": state["checkpoint_digest"],
        "state": str(state_path),
        "output": str(output),
        "total_source_records": len(records),
        "new_total": len(all_new),
        "new_in_batch": len(visible_items),
        "items": visible_items,
    }
    decision_payload = {
        "version": VERSION,
        "batch_token": token,
        "decisions": [
            {
                "id": item["id"],
                "fingerprint": item["fingerprint"],
                "decision": None,
                "scope": None,
                "reason": None,
                "matched_topics": [],
                "evidence_basis": None,
            }
            for item in visible_items
        ],
    }
    atomic_write_json(pending_path, pending_payload)
    atomic_write_json(decisions_path, decision_payload)
    summary = current_counts(records, state)
    summary.update(
        {
            "total_source_records": len(records),
            "new_total": len(all_new),
            "new_in_batch": len(visible_items),
            "pending_file": str(pending_path),
            "decisions_file": str(decisions_path),
        }
    )
    print(json.dumps(summary, ensure_ascii=False))


def validate_decisions(
    pending: dict[str, Any], decisions_payload: dict[str, Any]
) -> list[dict[str, Any]]:
    if pending.get("version") != VERSION or decisions_payload.get("version") != VERSION:
        raise FilterError("pending 或 decisions 版本不受支持")
    if decisions_payload.get("batch_token") != pending.get("batch_token"):
        raise FilterError("decisions 的 batch_token 与 pending 不一致")
    pending_items = pending.get("items")
    decisions = decisions_payload.get("decisions")
    if not isinstance(pending_items, list) or not isinstance(decisions, list):
        raise FilterError("pending.items 和 decisions.decisions 必须是数组")
    expected = {
        item.get("id"): item.get("fingerprint")
        for item in pending_items
        if isinstance(item, dict)
    }
    if len(expected) != len(pending_items):
        raise FilterError("pending 中存在无效或重复 id")
    if len(decisions) != len(expected):
        raise FilterError("decisions 必须完整覆盖当前批次，且不能添加额外条目")
    validated: list[dict[str, Any]] = []
    seen: set[str] = set()
    for index, decision_item in enumerate(decisions, start=1):
        if not isinstance(decision_item, dict):
            raise FilterError(f"第 {index} 条 decision 不是对象")
        item_id = decision_item.get("id")
        if not isinstance(item_id, str) or item_id not in expected or item_id in seen:
            raise FilterError(f"第 {index} 条 decision 的 id 无效或重复")
        seen.add(item_id)
        if decision_item.get("fingerprint") != expected[item_id]:
            raise FilterError(f"decision 指纹不匹配: {item_id}")
        decision = decision_item.get("decision")
        scope = decision_item.get("scope")
        evidence_basis = decision_item.get("evidence_basis")
        reason = decision_item.get("reason")
        topics = decision_item.get("matched_topics")
        if decision not in DECISIONS:
            raise FilterError(f"decision 值无效: {item_id}")
        if scope not in SCOPES:
            raise FilterError(f"scope 值无效: {item_id}")
        if decision == "review" and scope != "borderline":
            raise FilterError(f"review 必须使用 borderline scope: {item_id}")
        if decision == "excluded" and scope != "out-of-scope":
            raise FilterError(f"excluded 必须使用 out-of-scope scope: {item_id}")
        if decision == "relevant" and scope in {"borderline", "out-of-scope"}:
            raise FilterError(f"relevant 使用了不兼容的 scope: {item_id}")
        if evidence_basis not in EVIDENCE_BASES:
            raise FilterError(f"evidence_basis 值无效: {item_id}")
        if not isinstance(reason, str) or not reason.strip():
            raise FilterError(f"reason 不能为空: {item_id}")
        if not isinstance(topics, list) or any(not isinstance(topic, str) for topic in topics):
            raise FilterError(f"matched_topics 必须是字符串数组: {item_id}")
        validated.append(
            {
                "id": item_id,
                "fingerprint": expected[item_id],
                "decision": decision,
                "scope": scope,
                "reason": reason.strip(),
                "matched_topics": [topic.strip() for topic in topics if topic.strip()],
                "evidence_basis": evidence_basis,
            }
        )
    return validated


def style_sheet(sheet: Any) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:U{max(sheet.max_row, 1)}"
    widths = {
        "A": 45,
        "B": 32,
        "C": 10,
        "D": 30,
        "E": 28,
        "F": 20,
        "G": 45,
        "H": 45,
        "I": 60,
        "J": 24,
        "K": 30,
        "L": 20,
        "M": 20,
        "N": 12,
        "O": 55,
        "P": 12,
        "Q": 28,
        "R": 55,
        "S": 35,
        "T": 24,
        "U": 24,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in (12, 13, 21):
            row[column - 1].number_format = "yyyy-mm-dd hh:mm:ss"
        for column in (7, 8):
            cell: Cell = row[column - 1]
            if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"


def append_filtered_workbook(
    path: Path,
    current: dict[str, dict[str, Any]],
    decisions: list[dict[str, Any]],
    classified_at: str,
) -> int:
    headers = SOURCE_HEADERS + OUTPUT_META_HEADERS
    if path.exists():
        try:
            workbook = load_workbook(path)
        except Exception as exc:
            raise FilterError(f"无法读取筛选 Excel: {exc}") from exc
        if workbook.sheetnames != ["Papers"]:
            workbook.close()
            raise FilterError("筛选 Excel 必须且只能包含 Papers 工作表")
        sheet = workbook["Papers"]
        actual = [sheet.cell(1, index).value for index in range(1, len(headers) + 1)]
        if actual != headers:
            workbook.close()
            raise FilterError("筛选 Excel 表头不符合当前 Skill 版本")
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Papers"
        sheet.append(headers)
    existing_ids = {
        str(value).strip()
        for (value,) in sheet.iter_rows(
            min_row=2,
            min_col=SOURCE_HEADERS.index("dedup_key") + 1,
            max_col=SOURCE_HEADERS.index("dedup_key") + 1,
            values_only=True,
        )
        if str(value or "").strip()
    }
    added = 0
    for decision in decisions:
        if decision["decision"] != "relevant":
            continue
        record = current[decision["id"]]
        if decision["id"] in existing_ids:
            continue
        values = record["values"]
        sheet.append(
            [values.get(header) for header in SOURCE_HEADERS]
            + [
                record["source_row"],
                decision["scope"],
                decision["reason"],
                "; ".join(decision["matched_topics"]),
                decision["evidence_basis"],
                classified_at,
            ]
        )
        existing_ids.add(decision["id"])
        added += 1
    style_sheet(sheet)
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{path.stem}-", suffix=".xlsx", dir=path.parent, delete=False
        ) as handle:
            temporary = Path(handle.name)
        workbook.save(temporary)
        check = load_workbook(temporary, read_only=True, data_only=False)
        try:
            if check.sheetnames != ["Papers"]:
                raise FilterError("筛选 Excel 临时文件工作表验证失败")
            check_sheet = check["Papers"]
            check_header = next(check_sheet.iter_rows(max_row=1, values_only=True), ())
            if list(check_header[: len(headers)]) != headers:
                raise FilterError("筛选 Excel Papers 表头验证失败")
        finally:
            check.close()
        try:
            os.replace(temporary, path)
        except PermissionError as exc:
            raise FilterError(
                f"无法替换筛选 Excel，文件可能正被占用: {path}"
            ) from exc
        temporary = None
    finally:
        workbook.close()
        if temporary is not None:
            temporary.unlink(missing_ok=True)
    return added


def command_apply(args: argparse.Namespace) -> None:
    pending_path = args.pending.resolve()
    decisions_path = args.decisions.resolve()
    pending = load_json(pending_path)
    decisions_payload = load_json(decisions_path)
    decisions = validate_decisions(pending, decisions_payload)
    source = Path(str(pending.get("source") or ""))
    state_path = Path(str(pending.get("state") or ""))
    output = Path(str(pending.get("output") or ""))
    if not source.is_absolute() or not state_path.is_absolute() or not output.is_absolute():
        raise FilterError("pending 中的 source/state/output 必须是绝对路径")
    if sha256_file(source) != pending.get("source_snapshot"):
        raise FilterError("源 Excel 在 prepare 后已变化；请重新 prepare 并重新判定")
    records = read_source(source)
    current = {record["id"]: record for record in records}
    for decision in decisions:
        record = current.get(decision["id"])
        if record is None or record["fingerprint"] != decision["fingerprint"]:
            raise FilterError(f"源记录在 prepare 后已变化: {decision['id']}")
    old_state = load_state(state_path)
    validate_checkpoint(records, old_state)
    if old_state["last_source_row"] != pending.get("checkpoint_before"):
        raise FilterError("checkpoint 已被另一批次推进；请重新 prepare")
    if old_state["checkpoint_digest"] != pending.get("checkpoint_digest_before"):
        raise FilterError("checkpoint 摘要与 prepare 时不一致；请重新 prepare")
    new_records = dict(old_state["records"])
    classified_at = datetime.now().astimezone().isoformat(timespec="seconds")
    for decision in decisions:
        new_records[decision["id"]] = {
            "fingerprint": decision["fingerprint"],
            "decision": decision["decision"],
            "scope": decision["scope"],
            "reason": decision["reason"],
            "matched_topics": decision["matched_topics"],
            "evidence_basis": decision["evidence_basis"],
            "classified_at": classified_at,
            "source_row": current[decision["id"]]["source_row"],
        }
    new_last_source_row = (
        max(current[decision["id"]]["source_row"] for decision in decisions)
        if decisions
        else old_state["last_source_row"]
    )
    new_state = {
        "version": VERSION,
        "source": str(source),
        "updated_at": classified_at,
        "last_source_row": new_last_source_row,
        "checkpoint_digest": prefix_digest(records, new_last_source_row),
        "records": new_records,
    }
    added_to_filtered = append_filtered_workbook(
        output, current, decisions, classified_at
    )
    atomic_write_json(state_path, new_state)
    pending_path.unlink(missing_ok=True)
    decisions_path.unlink(missing_ok=True)
    summary = current_counts(records, new_state)
    summary.update(
        {
            "processed_new": len(decisions),
            "added_to_filtered": added_to_filtered,
            "output": str(output),
            "state": str(state_path),
        }
    )
    print(json.dumps(summary, ensure_ascii=False))


def command_status(args: argparse.Namespace) -> None:
    source = args.source.resolve()
    records = read_source(source)
    state = load_state(args.state.resolve())
    validate_checkpoint(records, state)
    summary = current_counts(records, state)
    summary["total_source_records"] = len(records)
    print(json.dumps(summary, ensure_ascii=False))


def command_reset(args: argparse.Namespace) -> None:
    state_path = args.state.resolve()
    if state_path.exists():
        timestamp = datetime.now().astimezone().strftime("%Y%m%dT%H%M%S%z")
        backup_path = state_path.with_name(f"{state_path.name}.bak-{timestamp}")
        os.replace(state_path, backup_path)
        print(
            json.dumps(
                {"reset": True, "state": str(state_path), "backup": str(backup_path)},
                ensure_ascii=False,
            )
        )
    else:
        print(json.dumps({"reset": False, "state": str(state_path)}, ensure_ascii=False))


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Incrementally maintain a DOA/array-relevant Excel shortlist."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    prepare = subparsers.add_parser("prepare", help="Prepare the next newly appended batch.")
    prepare.add_argument("--source", type=Path, required=True)
    prepare.add_argument("--state", type=Path, required=True)
    prepare.add_argument("--pending", type=Path, required=True)
    prepare.add_argument("--decisions", type=Path, required=True)
    prepare.add_argument("--output", type=Path, required=True)
    prepare.add_argument("--limit", type=int, default=30)
    prepare.set_defaults(handler=command_prepare)

    apply_command = subparsers.add_parser("apply", help="Validate and apply one decision batch.")
    apply_command.add_argument("--pending", type=Path, required=True)
    apply_command.add_argument("--decisions", type=Path, required=True)
    apply_command.set_defaults(handler=command_apply)

    status = subparsers.add_parser("status", help="Show current incremental classification counts.")
    status.add_argument("--source", type=Path, required=True)
    status.add_argument("--state", type=Path, required=True)
    status.set_defaults(handler=command_status)

    reset = subparsers.add_parser("reset", help="Archive the incremental checkpoint.")
    reset.add_argument("--state", type=Path, required=True)
    reset.set_defaults(handler=command_reset)
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    if getattr(args, "limit", 1) < 1:
        parser.error("--limit 必须大于 0")
    try:
        args.handler(args)
    except FilterError as exc:
        parser.exit(1, f"错误: {exc}\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
