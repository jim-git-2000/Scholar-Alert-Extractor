from datetime import datetime

from openpyxl import load_workbook

from scholar_alerts.excel_store import HEADERS, ExcelStore


def test_create_format_and_hyperlinks(tmp_path, paper_factory):
    path = tmp_path / "nested" / "papers.xlsx"
    result = ExcelStore(path).apply([paper_factory()])
    assert result.added == 1
    assert result.duplicates == 0
    workbook = load_workbook(path)
    sheet = workbook["Papers"]
    assert [cell.value for cell in sheet[1]] == HEADERS
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:O2"
    assert sheet["G2"].hyperlink.target == "https://ieeexplore.ieee.org/document/12345678"
    assert sheet["L2"].number_format == "yyyy-mm-dd hh:mm:ss"
    workbook.close()


def test_duplicate_updates_managed_values_and_fills_blanks(tmp_path, paper_factory):
    path = tmp_path / "papers.xlsx"
    store = ExcelStore(path)
    original = paper_factory(
        publication=None,
        pdf_url=None,
        source="google_scholar",
        alert_name="DOA",
    )
    store.apply([original], now=datetime(2026, 7, 1, 9, 0))
    duplicate = paper_factory(
        publication="User Curated Journal",
        pdf_url="https://example.org/paper.pdf",
        source="ieee_author_alert",
        alert_name="Alice Zhang",
    )
    result = store.apply([duplicate], now=datetime(2026, 7, 2, 9, 0))
    assert result.added == 0
    assert result.duplicates == 1
    workbook = load_workbook(path)
    sheet = workbook["Papers"]
    assert sheet.max_row == 2
    assert sheet["D2"].value == "User Curated Journal"
    assert sheet["H2"].value == "https://example.org/paper.pdf"
    assert set(sheet["J2"].value.split("; ")) == {"google_scholar", "ieee_author_alert"}
    assert set(sheet["K2"].value.split("; ")) == {"DOA", "Alice Zhang"}
    assert sheet["N2"].value == 2
    assert sheet["M2"].value == datetime(2026, 7, 2, 9, 0)
    workbook.close()


def test_nonempty_manual_field_is_not_overwritten(tmp_path, paper_factory):
    path = tmp_path / "papers.xlsx"
    store = ExcelStore(path)
    store.apply([paper_factory(publication="Original")])
    workbook = load_workbook(path)
    workbook["Papers"]["D2"] = "Manually corrected publication"
    workbook.save(path)
    workbook.close()
    store.apply([paper_factory(publication="Incoming value")])
    workbook = load_workbook(path)
    assert workbook["Papers"]["D2"].value == "Manually corrected publication"
    workbook.close()


def test_preview_does_not_create_or_modify_file(tmp_path, paper_factory):
    path = tmp_path / "papers.xlsx"
    result = ExcelStore(path).preview([paper_factory()])
    assert result.added == 1
    assert not path.exists()


def test_cross_source_url_or_title_year_matches_existing(tmp_path, paper_factory):
    path = tmp_path / "papers.xlsx"
    store = ExcelStore(path)
    store.apply([paper_factory(doi=None, ieee_document_id=None)])
    result = store.apply(
        [
            paper_factory(
                doi=None,
                ieee_document_id=None,
                source="google_scholar",
                paper_url="https://ieeexplore.ieee.org/document/12345678?utm_source=email",
            )
        ]
    )
    assert result.duplicates == 1
    assert store.count() == 1
