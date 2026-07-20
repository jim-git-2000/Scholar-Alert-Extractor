from __future__ import annotations

import os
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from scholar_alerts.dedup import candidate_keys, dedup_key, matching_key
from scholar_alerts.models import MergeResult, Paper

HEADERS = [
    "title",
    "authors",
    "year",
    "publication",
    "doi",
    "ieee_document_id",
    "paper_url",
    "pdf_url",
    "snippet",
    "sources",
    "alert_names",
    "first_seen_at",
    "last_seen_at",
    "seen_count",
    "dedup_key",
]
_OPTIONAL_FIELDS = (
    "authors",
    "year",
    "publication",
    "doi",
    "ieee_document_id",
    "paper_url",
    "pdf_url",
    "snippet",
)


class ExcelStoreError(RuntimeError):
    code = "excel_save_failed"


class ExcelValidationError(ExcelStoreError):
    code = "excel_validation_failed"


class ExcelLockedError(ExcelStoreError):
    code = "excel_locked"


def _naive(value: datetime) -> datetime:
    return value.astimezone().replace(tzinfo=None) if value.tzinfo else value.replace(tzinfo=None)


def _split_values(value: Any) -> list[str]:
    return [item.strip() for item in str(value or "").split(";") if item.strip()]


def _join_values(existing: Any, incoming: str | None) -> str | None:
    values = _split_values(existing)
    for value in _split_values(incoming):
        if value not in values:
            values.append(value)
    return "; ".join(values) or None


def _paper_from_row(values: dict[str, Any]) -> Paper:
    authors = [item.strip() for item in str(values.get("authors") or "").split(";") if item.strip()]
    year_value = values.get("year")
    try:
        year = int(year_value) if year_value not in (None, "") else None
    except (TypeError, ValueError):
        year = None
    return Paper(
        title=str(values.get("title") or ""),
        authors=authors,
        year=year,
        publication=values.get("publication") or None,
        doi=values.get("doi") or None,
        ieee_document_id=str(values.get("ieee_document_id") or "") or None,
        paper_url=values.get("paper_url") or None,
        pdf_url=values.get("pdf_url") or None,
        snippet=values.get("snippet") or None,
        source=str(values.get("sources") or ""),
        alert_name=values.get("alert_names") or None,
        received_at=values.get("first_seen_at") or datetime.now(),
        message_id=None,
        email_uid=0,
    )


class ExcelStore:
    def __init__(self, path: Path):
        self.path = path

    def count(self) -> int:
        if not self.path.exists():
            return 0
        workbook, sheet = self._load()
        try:
            return max(sheet.max_row - 1, 0)
        finally:
            workbook.close()

    def preview(self, papers: list[Paper], *, now: datetime | None = None) -> MergeResult:
        workbook, sheet = self._load()
        try:
            return self._merge(sheet, papers, now=now or datetime.now())
        finally:
            workbook.close()

    def apply(self, papers: list[Paper], *, now: datetime | None = None) -> MergeResult:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        workbook, sheet = self._load()
        temporary_path: Path | None = None
        try:
            result = self._merge(sheet, papers, now=now or datetime.now())
            self._format(sheet)
            with tempfile.NamedTemporaryFile(
                prefix=f".{self.path.stem}-", suffix=".xlsx", dir=self.path.parent, delete=False
            ) as handle:
                temporary_path = Path(handle.name)
            workbook.save(temporary_path)
            self._validate_file(
                temporary_path,
                expected_rows=result.total,
                expected_keys=result.changed_keys,
            )
            try:
                os.replace(temporary_path, self.path)
            except PermissionError as exc:
                raise ExcelLockedError(f"Excel 文件可能正被占用: {self.path}") from exc
            temporary_path = None
            self._validate_file(
                self.path,
                expected_rows=result.total,
                expected_keys=result.changed_keys,
            )
            return result
        except ExcelStoreError:
            raise
        except PermissionError as exc:
            raise ExcelLockedError(f"无法写入 Excel: {self.path}") from exc
        except OSError as exc:
            raise ExcelStoreError(f"保存 Excel 失败: {exc}") from exc
        finally:
            workbook.close()
            if temporary_path is not None:
                temporary_path.unlink(missing_ok=True)

    def _load(self) -> tuple[Workbook, Worksheet]:
        if self.path.exists():
            try:
                workbook = load_workbook(self.path)
            except PermissionError as exc:
                raise ExcelLockedError(f"Excel 文件可能正被占用: {self.path}") from exc
            except Exception as exc:
                raise ExcelValidationError(f"无法读取 Excel: {exc}") from exc
            if "Papers" not in workbook.sheetnames:
                workbook.close()
                raise ExcelValidationError("Excel 缺少 Papers 工作表")
            sheet = workbook["Papers"]
            self._validate_sheet(sheet)
            return workbook, sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Papers"
        sheet.append(HEADERS)
        self._format(sheet)
        return workbook, sheet

    @staticmethod
    def _validate_sheet(sheet: Worksheet) -> None:
        actual = [sheet.cell(1, column).value for column in range(1, len(HEADERS) + 1)]
        if actual != HEADERS:
            raise ExcelValidationError("Excel 表头与固定列定义不一致")

    def _validate_file(
        self, path: Path, *, expected_rows: int, expected_keys: list[str]
    ) -> None:
        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
            try:
                if "Papers" not in workbook.sheetnames:
                    raise ExcelValidationError("保存后的 Excel 缺少 Papers 工作表")
                sheet = workbook["Papers"]
                self._validate_sheet(sheet)
                actual_rows = max(sheet.max_row - 1, 0)
                if actual_rows != expected_rows:
                    raise ExcelValidationError(
                        
                            f"保存后的 Excel 行数异常: 预期 {expected_rows}，"
                            f"实际 {actual_rows}"
                        
                    )
                saved_keys = {
                    str(sheet.cell(row, len(HEADERS)).value or "")
                    for row in range(2, sheet.max_row + 1)
                }
                missing_keys = set(expected_keys) - saved_keys
                if missing_keys:
                    raise ExcelValidationError(
                        f"保存后的 Excel 缺少 {len(missing_keys)} 条变更记录"
                    )
            finally:
                workbook.close()
        except ExcelStoreError:
            raise
        except Exception as exc:
            raise ExcelValidationError(f"无法验证 Excel: {exc}") from exc

    def _merge(self, sheet: Worksheet, papers: list[Paper], *, now: datetime) -> MergeResult:
        index: dict[str, set[int]] = {}
        row_values: dict[int, dict[str, Any]] = {}
        for row_number in range(2, sheet.max_row + 1):
            values = {
                header: sheet.cell(row_number, column).value
                for column, header in enumerate(HEADERS, start=1)
            }
            row_values[row_number] = values
            existing_paper = _paper_from_row(values)
            for key in candidate_keys(existing_paper):
                index.setdefault(key, set()).add(row_number)
            stored_key = str(values.get("dedup_key") or "").strip()
            if stored_key:
                index.setdefault(stored_key, set()).add(row_number)

        added = 0
        duplicates = 0
        changed_keys: list[str] = []
        timestamp = _naive(now).replace(microsecond=0)
        for paper in papers:
            keys = candidate_keys(paper)
            if not keys:
                raise ExcelValidationError(f"论文无法生成去重键: {paper.title}")
            candidate_rows = {
                row_number for key in keys for row_number in index.get(key, set())
            }
            row_number = next(
                (
                    candidate_row
                    for candidate_row in sorted(candidate_rows)
                    if matching_key(paper, _paper_from_row(row_values[candidate_row]))
                ),
                None,
            )
            if row_number is None:
                key = dedup_key(paper)
                values = {
                    "title": paper.title,
                    "authors": "; ".join(paper.authors) or None,
                    "year": paper.year,
                    "publication": paper.publication,
                    "doi": paper.doi,
                    "ieee_document_id": paper.ieee_document_id,
                    "paper_url": paper.paper_url,
                    "pdf_url": paper.pdf_url,
                    "snippet": paper.snippet,
                    "sources": paper.source,
                    "alert_names": paper.alert_name,
                    "first_seen_at": timestamp,
                    "last_seen_at": timestamp,
                    "seen_count": 1,
                    "dedup_key": key,
                }
                sheet.append([values[header] for header in HEADERS])
                row_number = sheet.max_row
                row_values[row_number] = values
                added += 1
            else:
                values = row_values[row_number]
                incoming = {
                    "authors": "; ".join(paper.authors) or None,
                    "year": paper.year,
                    "publication": paper.publication,
                    "doi": paper.doi,
                    "ieee_document_id": paper.ieee_document_id,
                    "paper_url": paper.paper_url,
                    "pdf_url": paper.pdf_url,
                    "snippet": paper.snippet,
                }
                for field_name in _OPTIONAL_FIELDS:
                    if (
                        values.get(field_name) in (None, "")
                        and incoming[field_name] not in (None, "")
                    ):
                        values[field_name] = incoming[field_name]
                values["sources"] = _join_values(values.get("sources"), paper.source)
                values["alert_names"] = _join_values(values.get("alert_names"), paper.alert_name)
                values["last_seen_at"] = timestamp
                try:
                    values["seen_count"] = int(values.get("seen_count") or 0) + 1
                except (TypeError, ValueError):
                    values["seen_count"] = 1
                for column, header in enumerate(HEADERS, start=1):
                    sheet.cell(row_number, column).value = values.get(header)
                duplicates += 1

            current_paper = _paper_from_row(row_values[row_number])
            current_key = dedup_key(current_paper)
            row_values[row_number]["dedup_key"] = current_key
            sheet.cell(row_number, HEADERS.index("dedup_key") + 1).value = current_key
            changed_key = current_key
            changed_keys.append(changed_key)
            for key in set(keys + candidate_keys(current_paper)):
                index.setdefault(key, set()).add(row_number)

        return MergeResult(
            added=added,
            duplicates=duplicates,
            total=max(sheet.max_row - 1, 0),
            changed_keys=changed_keys,
        )

    @staticmethod
    def _format(sheet: Worksheet) -> None:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:O{max(sheet.max_row, 1)}"
        widths = {
            "A": 45,
            "B": 32,
            "C": 10,
            "D": 30,
            "E": 28,
            "F": 20,
            "G": 45,
            "H": 45,
            "I": 60,
            "J": 24,
            "K": 30,
            "L": 20,
            "M": 20,
            "N": 12,
            "O": 55,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            for column in (12, 13):
                row[column - 1].number_format = "yyyy-mm-dd hh:mm:ss"
            for column in (7, 8):
                cell: Cell = row[column - 1]
                if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
